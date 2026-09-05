from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from admissions.models import Santri
from core.models import WebsiteSettings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from io import BytesIO


def get_all_santri_data(santri):
    """Get semua data santri dalam format dictionary untuk Excel"""
    website_settings = WebsiteSettings.load()
    
    # Format tanggal
    def format_date(date_obj):
        if date_obj:
            return date_obj.strftime('%d/%m/%Y')
        return ''
    
    # Format tanggal lengkap
    def format_date_long(date_obj):
        if date_obj:
            months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                     'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
            return f"{date_obj.day} {months[date_obj.month - 1]} {date_obj.year}"
        return ''
    
    # Alamat lengkap
    alamat_parts = filter(None, [
        santri.alamat,
        santri.desa,
        santri.kecamatan,
        santri.kabupaten,
        santri.provinsi,
        santri.kode_pos
    ])
    alamat_lengkap = ', '.join(alamat_parts) if alamat_parts else ''
    
    return {
        # Identitas Santri (Penting)
        'No_Pendaftaran': f"reg_{santri.id:07d}",
        'NISN': santri.nisn or '',
        'Nama_Lengkap': santri.nama_lengkap or '',
        'Nama_Panggilan': santri.nama_panggilan or '',
        'Tempat_Lahir': santri.tempat_lahir or '',
        'Tanggal_Lahir': format_date(santri.tanggal_lahir),
        'Tanggal_Lahir_Lengkap': format_date_long(santri.tanggal_lahir),
        'Jenis_Kelamin': santri.get_jenis_kelamin_display() if santri.jenis_kelamin else '',
        'Agama': santri.agama or 'Islam',
        'Kewarganegaraan': santri.kewarganegaraan or 'WNI',
        
        # Alamat Santri (Penting)
        'Alamat': santri.alamat or '',
        'Desa': santri.desa or '',
        'Kecamatan': santri.kecamatan or '',
        'Kabupaten': santri.kabupaten or '',
        'Provinsi': santri.provinsi or '',
        'Kode_Pos': santri.kode_pos or '',
        'Alamat_Lengkap': alamat_lengkap,
        
        # Data Ayah (Penting)
        'Nama_Ayah': santri.nama_ayah or '',
        'NIK_Ayah': santri.nik_ayah or '',
        'Tempat_Lahir_Ayah': santri.tempat_lahir_ayah or '',
        'Tanggal_Lahir_Ayah': format_date(santri.tanggal_lahir_ayah),
        'Tanggal_Lahir_Ayah_Lengkap': format_date_long(santri.tanggal_lahir_ayah),
        'Agama_Ayah': santri.agama_ayah or 'Islam',
        'Kewarganegaraan_Ayah': santri.kewarganegaraan_ayah or 'WNI',
        'Pendidikan_Ayah': santri.pendidikan_ayah or '',
        'Pekerjaan_Ayah': santri.pekerjaan_ayah or '',
        'No_HP_Ayah': santri.no_hp_ayah or '',
        'Status_Ayah': santri.status_ayah or 'HIDUP',
        
        # Data Ibu (Penting)
        'Nama_Ibu': santri.nama_ibu or '',
        'NIK_Ibu': santri.nik_ibu or '',
        'Tempat_Lahir_Ibu': santri.tempat_lahir_ibu or '',
        'Tanggal_Lahir_Ibu': format_date(santri.tanggal_lahir_ibu),
        'Tanggal_Lahir_Ibu_Lengkap': format_date_long(santri.tanggal_lahir_ibu),
        'Agama_Ibu': santri.agama_ibu or 'Islam',
        'Kewarganegaraan_Ibu': santri.kewarganegaraan_ibu or 'WNI',
        'Pendidikan_Ibu': santri.pendidikan_ibu or '',
        'Pekerjaan_Ibu': santri.pekerjaan_ibu or '',
        'No_HP_Ibu': santri.no_hp_ibu or '',
        'Status_Ibu': santri.status_ibu or 'HIDUP',
        'Alamat_Orangtua': santri.alamat_orangtua or alamat_lengkap,
        
        # Data Sekolah (Penting)
        'Asal_Sekolah': santri.asal_sekolah or '',
        'NPSN_Sekolah': santri.npsn_sekolah or '',
        'Kelas_Terakhir': santri.kelas_terakhir or '',
        'Tahun_Lulus': santri.tahun_lulus or '',
        'No_Ijazah': santri.no_ijazah or '',
        'Kelas_Diterima': santri.kelas_diterima or '',
        'Tanggal_Diterima': format_date(santri.tanggal_diterima),
        'Tanggal_Diterima_Lengkap': format_date_long(santri.tanggal_diterima),
    }




@login_required
@staff_member_required
def tutorial(request):
    """Tutorial Mail Merge"""
    return render(request, 'admin-panel/documents/tutorial.html')


@login_required
@staff_member_required
def document_list(request):
    """List semua santri untuk generate Excel"""
    from django.db.models import Q
    search = request.GET.get('search', '')
    santri_list = Santri.objects.all().order_by('-created_at')
    
    if search:
        santri_list = santri_list.filter(
            Q(nama_lengkap__icontains=search) |
            Q(nisn__icontains=search) |
            Q(email__icontains=search)
        )
    
    context = {
        'santri_list': santri_list,
        'search': search,
    }
    return render(request, 'admin-panel/documents/list.html', context)


@login_required
@staff_member_required
def generate_excel(request, santri_id=None):
    """Generate Excel untuk Mail Merge Word"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Santri"
        
        # Get santri data
        if santri_id:
            # Single santri
            santri = get_object_or_404(Santri, id=santri_id)
            santri_list = [santri]
            filename = f"Data_Santri_{santri.nisn}_{santri.nama_lengkap.replace(' ', '_')}.xlsx"
        else:
            # All santri
            santri_list = Santri.objects.all().order_by('nama_lengkap')
            filename = f"Data_Semua_Santri_{timezone.now().strftime('%Y%m%d')}.xlsx"
        
        if not santri_list:
            messages.error(request, 'Tidak ada data santri untuk diekspor')
            return redirect('documents:list')
        
        # Get first santri data untuk menentukan kolom
        first_santri_data = get_all_santri_data(santri_list[0])
        columns = list(first_santri_data.keys())
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write header (format untuk Word Mail Merge - tanpa spasi, gunakan underscore)
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
            # Auto adjust column width
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(column_title) + 2, 15)
        
        # Set print area dan page setup untuk Excel yang lebih rapi
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Write data rows
        for row_num, santri in enumerate(santri_list, 2):
            santri_data = get_all_santri_data(santri)
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = santri_data.get(column_title, '')
                cell.border = border_style
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('documents:list')


