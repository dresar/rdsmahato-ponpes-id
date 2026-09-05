"""
Utilities untuk Data Management: Export, Import, Bulk Actions
"""
import csv
import io
import os
from datetime import datetime
from django.http import HttpResponse
from django.db import models
from django.utils import timezone

# Fix untuk error numpy: Set environment variable sebelum import
# Ini harus dilakukan SEBELUM import numpy atau library yang menggunakan numpy
os.environ.setdefault('OPENBLAS_NUM_THREADS', '1')
os.environ.setdefault('MKL_NUM_THREADS', '1')
os.environ.setdefault('NUMEXPR_NUM_THREADS', '1')
os.environ.setdefault('OMP_NUM_THREADS', '1')

# Import reportlab terlebih dahulu (tidak menggunakan numpy)
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Lazy import untuk openpyxl dan pandas untuk menghindari konflik numpy
# Import dilakukan di dalam fungsi yang membutuhkan untuk menghindari konflik
_openpyxl_available = False
_pandas_available = False

def _import_openpyxl():
    """Import openpyxl dengan error handling"""
    global _openpyxl_available
    if _openpyxl_available:
        return True
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        globals()['Workbook'] = Workbook
        globals()['Font'] = Font
        globals()['Alignment'] = Alignment
        globals()['PatternFill'] = PatternFill
        _openpyxl_available = True
        return True
    except Exception as e:
        return False

def _import_pandas():
    """Import pandas dengan error handling"""
    global _pandas_available
    if _pandas_available:
        return True
    try:
        import pandas as pd
        globals()['pd'] = pd
        _pandas_available = True
        return True
    except Exception as e:
        return False


class DataExporter:
    """Utility class untuk export data ke berbagai format"""
    
    @staticmethod
    def export_to_excel(queryset, model_name, fields, headers=None):
        """Export queryset ke Excel"""
        if not _import_openpyxl():
            raise ImportError("openpyxl tidak terinstall atau terjadi error. Install dengan: pip install openpyxl")
        wb = Workbook()
        ws = wb.active
        ws.title = model_name
        
        # Headers
        if headers is None:
            headers = [field.replace('_', ' ').title() for field in fields]
        
        # Style untuk header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Tulis header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Tulis data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                value = getattr(obj, field, '')
                
                # Handle datetime fields
                if isinstance(value, datetime):
                    # Convert timezone-aware datetime ke local time jika perlu
                    if timezone.is_aware(value):
                        value = timezone.localtime(value)
                    # Format tanggal untuk Excel: DD/MM/YYYY HH:MM
                    value = value.strftime('%d/%m/%Y %H:%M')
                # Handle date fields
                elif hasattr(value, 'strftime') and not isinstance(value, (str, int, float, bool, type(None))):
                    # Handle date objects
                    try:
                        value = value.strftime('%d/%m/%Y')
                    except:
                        value = str(value)
                # Handle ForeignKey and other objects
                elif hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num in range(1, len(fields) + 1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Buat response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_to_csv(queryset, model_name, fields, headers=None):
        """Export queryset ke CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # BOM untuk Excel compatibility
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Headers
        if headers is None:
            headers = [field.replace('_', ' ').title() for field in fields]
        writer.writerow(headers)
        
        # Data
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, '')
                
                # Handle datetime fields
                if isinstance(value, datetime):
                    # Convert timezone-aware datetime ke local time jika perlu
                    if timezone.is_aware(value):
                        value = timezone.localtime(value)
                    # Format tanggal untuk CSV: DD/MM/YYYY HH:MM
                    value = value.strftime('%d/%m/%Y %H:%M')
                # Handle date fields
                elif hasattr(value, 'strftime') and not isinstance(value, (str, int, float, bool, type(None))):
                    # Handle date objects
                    try:
                        value = value.strftime('%d/%m/%Y')
                    except:
                        value = str(value)
                # Handle other objects
                elif hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                
                row.append(value)
            writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_to_pdf(queryset, model_name, fields, headers=None, title=None):
        """Export queryset ke PDF dengan orientasi landscape"""
        response = HttpResponse(content_type='application/pdf')
        filename = f"{model_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Buat PDF dengan orientasi landscape
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        if title is None:
            title = f"Data {model_name}"
        title_para = Paragraph(title, styles['Title'])
        elements.append(title_para)
        elements.append(Spacer(1, 0.15*inch))
        
        # Info
        export_time = timezone.localtime(timezone.now())
        info_text = f"Tanggal Export: {export_time.strftime('%d/%m/%Y %H:%M')}<br/>"
        info_text += f"Total Data: {queryset.count()}"
        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 0.15*inch))
        
        # Headers
        if headers is None:
            headers = [field.replace('_', ' ').title() for field in fields]
        
        # Data untuk tabel
        data = [headers]
        
        # Batasi data untuk PDF (max 200 rows karena landscape lebih lebar)
        queryset_limited = queryset[:200]
        for obj in queryset_limited:
            row = []
            for field in fields:
                value = getattr(obj, field, '')
                
                # Handle datetime fields (created_at, updated_at, tanggal_lahir, etc.)
                if isinstance(value, datetime):
                    # Convert timezone-aware datetime ke local time jika perlu
                    if timezone.is_aware(value):
                        value = timezone.localtime(value)
                    # Format tanggal lebih singkat untuk PDF: DD/MM/YYYY
                    value = value.strftime('%d/%m/%Y')
                # Handle date fields (tanggal_lahir, etc.)
                elif hasattr(value, 'strftime') and not isinstance(value, (str, int, float, bool, type(None))):
                    # Handle date objects
                    try:
                        value = value.strftime('%d/%m/%Y')
                    except:
                        value = str(value)
                # Handle other objects
                elif hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                
                # Truncate panjang teks - lebih panjang karena landscape dan kolom lebih sedikit
                if isinstance(value, str) and len(value) > 40:
                    value = value[:37] + '...'
                row.append(str(value))
            data.append(row)
        
        # Hitung lebar kolom (landscape A4 width = 11.69 inch)
        num_cols = len(headers)
        # Lebar kolom disesuaikan dengan jumlah kolom (lebih lebar jika kolom lebih sedikit)
        available_width = 11.69 * inch - 1.5 * inch  # Margin kiri-kanan
        col_width = available_width / num_cols
        
        # Set lebar kolom dengan proporsi yang lebih baik
        table_widths = [col_width] * num_cols
        
        # Buat tabel dengan lebar kolom yang disesuaikan
        table = Table(data, colWidths=table_widths, repeatRows=1)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),  # Warna hijau gelap
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header center
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),   # Data left
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return response


class DataImporter:
    """Utility class untuk import data dari Excel/CSV"""
    
    @staticmethod
    def import_from_excel(file, model_class, field_mapping, skip_rows=0):
        """Import data dari Excel file"""
        if not _import_pandas():
            raise ImportError("pandas tidak terinstall atau terjadi error. Install dengan: pip install pandas")
        try:
            df = pd.read_excel(file, skiprows=skip_rows)
            return DataImporter._import_from_dataframe(df, model_class, field_mapping)
        except Exception as e:
            raise Exception(f"Error membaca file Excel: {str(e)}")
    
    @staticmethod
    def import_from_csv(file, model_class, field_mapping, encoding='utf-8-sig', skip_rows=0):
        """Import data dari CSV file"""
        if not _import_pandas():
            raise ImportError("pandas tidak terinstall atau terjadi error. Install dengan: pip install pandas")
        try:
            df = pd.read_csv(file, encoding=encoding, skiprows=skip_rows)
            return DataImporter._import_from_dataframe(df, model_class, field_mapping)
        except Exception as e:
            raise Exception(f"Error membaca file CSV: {str(e)}")
    
    @staticmethod
    def _import_from_dataframe(df, model_class, field_mapping):
        """Import dari pandas DataFrame"""
        imported = 0
        errors = []
        
        # Normalize column names (lowercase, replace spaces with underscore)
        df.columns = df.columns.str.lower().str.replace(' ', '_').str.strip()
        
        for index, row in df.iterrows():
            try:
                data = {}
                for excel_col, model_field in field_mapping.items():
                    excel_col_normalized = excel_col.lower().replace(' ', '_').strip()
                    if excel_col_normalized in df.columns:
                        value = row[excel_col_normalized]
                        # Handle NaN
                        if pd.isna(value):
                            value = None
                        # Handle empty string
                        if isinstance(value, str) and value.strip() == '':
                            value = None
                        # Set default untuk agama jika kosong
                        if model_field == 'agama' and (value is None or value == ''):
                            value = 'Islam'
                        data[model_field] = value
                
                # Set default agama jika tidak ada di mapping atau kosong
                if 'agama' not in data or not data.get('agama'):
                    data['agama'] = 'Islam'
                
                # Buat atau update object
                obj, created = model_class.objects.get_or_create(**data)
                if created:
                    imported += 1
            except Exception as e:
                errors.append(f"Baris {index + 2}: {str(e)}")
        
        return {
            'imported': imported,
            'errors': errors,
            'total_rows': len(df)
        }


class BulkActionHandler:
    """Utility untuk handle bulk actions"""
    
    @staticmethod
    def bulk_delete(queryset, ids):
        """Bulk delete objects"""
        deleted_count = 0
        errors = []
        
        for obj_id in ids:
            try:
                obj = queryset.get(id=obj_id)
                obj.delete()
                deleted_count += 1
            except Exception as e:
                errors.append(f"ID {obj_id}: {str(e)}")
        
        return {
            'deleted': deleted_count,
            'errors': errors,
            'total': len(ids)
        }
    
    @staticmethod
    def bulk_update(queryset, ids, update_data):
        """Bulk update objects"""
        updated_count = 0
        errors = []
        
        for obj_id in ids:
            try:
                obj = queryset.get(id=obj_id)
                for field, value in update_data.items():
                    setattr(obj, field, value)
                obj.save()
                updated_count += 1
            except Exception as e:
                errors.append(f"ID {obj_id}: {str(e)}")
        
        return {
            'updated': updated_count,
            'errors': errors,
            'total': len(ids)
        }

